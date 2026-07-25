from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY       = "1B2A4A"
LIGHT_GRAY = "F5F5F5"
WHITE      = "FFFFFF"
BORDER_CLR = "CCCCCC"

COLUMNS = [
    ("Job Title",  25),
    ("Company",    20),
    ("Location",   16),
    ("Type",       12),
    ("Posted",     12),
    ("ATS Score",  10),
    ("Tailored",   12),
    ("Gaps (missing keywords)", 48),
    ("Apply Link", 14),
    ("CV File",    14),
]

thin_border = Border(
    left=Side(style="thin", color=BORDER_CLR),
    right=Side(style="thin", color=BORDER_CLR),
    top=Side(style="thin", color=BORDER_CLR),
    bottom=Side(style="thin", color=BORDER_CLR),
)

header_font  = Font(name="Arial", size=11, bold=True, color=WHITE)
header_fill  = PatternFill("solid", fgColor=NAVY)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font    = Font(name="Arial", size=10, color="333333")
body_align   = Alignment(vertical="center", wrap_text=True)
link_font    = Font(name="Arial", size=10, color="1155CC", underline="single")

score_good_font = Font(name="Arial", size=10, bold=True, color="0F6E56")
score_ok_font   = Font(name="Arial", size=10, bold=True, color="854F0B")
score_bad_font  = Font(name="Arial", size=10, bold=True, color="993C1D")


def build_xlsx(job_results: list[dict], output_path: str) -> str:
    """Build the job-tracker spreadsheet.

    Each job dict may include a `cv_url` (a pre-signed Supabase URL) so the
    "CV File" hyperlink works when opened in Excel outside the app.
    """
    import os

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Results"

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font, cell.fill, cell.alignment, cell.border = (
            header_font, header_fill, header_align, thin_border)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 30

    for row_idx, job in enumerate(job_results, start=2):
        row_fill = PatternFill("solid", fgColor=WHITE if row_idx % 2 == 0 else LIGHT_GRAY)
        posted = (job.get("posted_at") or "")[:10]
        tailored = "Tailored" if job.get("tailored") else "Original"

        values = [
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", "") or "—",
            job.get("employment_type", ""),
            posted,
            job.get("ats_score", "N/A"),
            tailored,
            job.get("gaps", "") or "—",
            "Apply",
            "Download",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font, cell.alignment, cell.fill, cell.border = (
                body_font, body_align, row_fill, thin_border)

        # Apply-link hyperlink (col 9)
        if job.get("apply_link"):
            c = ws.cell(row=row_idx, column=9)
            c.hyperlink, c.value, c.font = job["apply_link"], "Apply", link_font

        # CV download hyperlink (col 10) — pre-signed Supabase URL
        if job.get("cv_url"):
            c = ws.cell(row=row_idx, column=10)
            c.hyperlink, c.value, c.font = job["cv_url"], "Download", link_font

        # ATS score colour (col 6)
        score = job.get("ats_score", "N/A")
        sc = ws.cell(row=row_idx, column=6)
        if isinstance(score, (int, float)):
            sc.value = f"{int(score)}%"
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.font = score_good_font if score >= 70 else score_ok_font if score >= 55 else score_bad_font

    ws.freeze_panes = "A2"
    dir_name = os.path.dirname(output_path)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)
    wb.save(output_path)
    return output_path
